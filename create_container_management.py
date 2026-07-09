#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
نظام إدارة الحاويات - Container Management System
يقوم بإنشاء ملف Excel احترافي شامل مع:
- لوحة تحكم (Dashboard)
- نظام إدخال البيانات
- قاعدة بيانات منظمة
- نظام تصاريح متقدم
- التقارير والشهادات
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    Protection, NamedStyle
)
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.chart import (
    BarChart, PieChart, LineChart, Reference
)
from datetime import datetime, timedelta
import random

# ============================================================================
# إعدادات التصميم (Design Settings)
# ============================================================================

COLORS = {
    'dark_bg': '0D1B2A',      # أسود/رمادي داكن جداً
    'header': '003D82',        # أزرق داكن عميق
    'accent': '0066CC',        # أزرق لازوردي
    'secondary': '1A4D99',     # أزرق فاتح
    'text_white': 'FFFFFF',    # أبيض
    'success': '27AE60',       # أخضر
    'warning': 'E67E22',       # برتقالي
    'danger': 'C0392B',        # أحمر
    'light_gray': 'ECF0F1',    # رمادي فاتح
}

COUNTRIES = ['السعودية', 'الإمارات', 'قطر', 'الكويت', 'البحرين', 'عمان']
COMPANIES = ['الراجحي', 'سابك', 'أرامكو', 'كهرباء السعودية', 'سوميتومو', 'الاتصالات']
CONTAINER_SIZES = ['20 قدم', '40 قدم', 'High Cube']
STATUS_OPTIONS = ['✅ يمكن إرسالها', '⏸️ محجوزة (انتظار تصريح)', '🔒 ممنوع الإرسال', '⚠️ موافقة مشروطة']

# ============================================================================
# البيانات التجريبية (Sample Data)
# ============================================================================

def generate_sample_containers():
    """توليد 15 حاوية تجريبية واقعية"""
    containers = []
    base_date = datetime.now()

    for i in range(1, 16):
        container = {
            'container_number': f'CMAU{random.randint(100000, 999999)}',
            'size': random.choice(CONTAINER_SIZES),
            'arrival_date': (base_date - timedelta(days=random.randint(1, 60))).strftime('%d/%m/%Y'),
            'company': random.choice(COMPANIES),
            'country': random.choice(COUNTRIES),
            'contents': random.choice(['مكائن', 'مواد خام', 'قطع غيار', 'منتجات', 'أجهزة']),
            'weight_tons': round(random.uniform(5, 25), 1),
            'notes': 'حاوية جديدة' if random.random() > 0.7 else 'مراجعة روتينية',
            'status': random.choice(STATUS_OPTIONS),
            'conditions': '',
            'approval_date': (base_date - timedelta(days=random.randint(1, 30))).strftime('%d/%m/%Y'),
        }

        # إضافة شروط إذا كانت الموافقة مشروطة
        if container['status'] == '⚠️ موافقة مشروطة':
            container['conditions'] = random.choice([
                'فحص إضافي مطلوب',
                'وثائق معلقة',
                'تصريح أمني',
                'معاينة طبية'
            ])

        containers.append(container)

    return containers

# ============================================================================
# إنشاء Workbook
# ============================================================================

def create_workbook():
    """إنشاء Workbook جديد"""
    wb = Workbook()
    wb.remove(wb.active)  # حذف الورقة الفارغة الافتراضية

    # تعيين الاتجاه RTL للمصنف
    wb.properties.language = 'ar'

    return wb

# ============================================================================
# دوال مساعدة للتنسيق
# ============================================================================

def style_header_cell(cell, text, color=COLORS['header']):
    """تنسيق خلية الرأس"""
    cell.value = text
    cell.font = Font(name='Calibri', size=12, bold=True, color=COLORS['text_white'])
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )
    return cell

def style_data_cell(cell, value, is_right=True):
    """تنسيق خلية بيانات"""
    cell.value = value
    cell.font = Font(name='Calibri', size=11, color='000000')
    cell.alignment = Alignment(
        horizontal='right' if is_right else 'center',
        vertical='center',
        wrap_text=True
    )
    return cell

def style_row(ws, row_num, is_header=False, color=None):
    """تنسيق صف كامل"""
    for cell in ws[row_num]:
        if cell.value is not None:
            if is_header:
                style_header_cell(cell, cell.value, color or COLORS['header'])
            else:
                # تلوين الصفوف بالتناوب
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                else:
                    alt_color = COLORS['light_gray'] if row_num % 2 == 0 else COLORS['text_white']
                    cell.fill = PatternFill(start_color=alt_color, end_color=alt_color, fill_type='solid')

# ============================================================================
# إنشاء الأوراق
# ============================================================================

def create_dashboard(wb, containers):
    """إنشاء لوحة التحكم - Dashboard"""
    ws = wb.create_sheet('📊 لوحة التحكم', 0)
    ws.right_to_left = True  # RTL
    ws.sheet_properties.tabColor = '0066CC'

    # تعيين الأعمدة
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    for col in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15

    # العنوان الرئيسي
    ws.row_dimensions[1].height = 35
    title_cell = ws['A1']
    title_cell.value = 'نظام إدارة الحاويات - لوحة التحكم'
    title_cell.font = Font(name='Calibri', size=20, bold=True, color=COLORS['text_white'])
    title_cell.fill = PatternFill(start_color=COLORS['dark_bg'], end_color=COLORS['dark_bg'], fill_type='solid')
    title_cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells('A1:F1')

    # KPI Cards
    ws.row_dimensions[3].height = 25
    kpi_row = 3
    kpi_labels = ['الحاويات الموجودة', 'حاويات 40 قدم', 'حاويات 20 قدم', 'المحجوزة']

    for idx, label in enumerate(kpi_labels):
        col = get_column_letter(6 - idx)  # من اليمين لليسار
        ws[f'{col}{kpi_row}'].value = label
        style_header_cell(ws[f'{col}{kpi_row}'], label, COLORS['secondary'])
        ws.row_dimensions[kpi_row].height = 30

    # البيانات - KPI values
    kpi_values = [
        len(containers),
        len([c for c in containers if c['size'] == '40 قدم']),
        len([c for c in containers if c['size'] == '20 قدم']),
        len([c for c in containers if 'محجوزة' in c['status']])
    ]

    ws.row_dimensions[4].height = 30
    for idx, value in enumerate(kpi_values):
        col = get_column_letter(6 - idx)
        ws[f'{col}4'].value = value
        ws[f'{col}4'].font = Font(name='Calibri', size=16, bold=True, color=COLORS['accent'])
        ws[f'{col}4'].fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type='solid')
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')

    # جدول البيانات الأخيرة
    ws.row_dimensions[6].height = 25
    table_headers = ['الحالة', 'البلد', 'الشركة', 'المحتويات', 'رقم الحاوية']
    for idx, header in enumerate(table_headers):
        col = get_column_letter(6 - idx)
        style_header_cell(ws[f'{col}6'], header)

    # البيانات - آخر 5 حاويات
    for row_idx, container in enumerate(containers[:5], start=7):
        ws.row_dimensions[row_idx].height = 20
        col_data = [container['status'], container['country'], container['company'],
                   container['contents'], container['container_number']]

        for col_idx, data in enumerate(col_data):
            col = get_column_letter(6 - col_idx)
            style_data_cell(ws[f'{col}{row_idx}'], data)

    return ws

def create_add_container_sheet(wb):
    """إنشاء ورقة إضافة حاوية قادمة"""
    ws = wb.create_sheet('📥 إضافة حاوية قادمة', 1)
    ws.right_to_left = True
    ws.sheet_properties.tabColor = '27AE60'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    # العنوان
    ws.row_dimensions[1].height = 30
    title = ws['A1']
    title.value = 'إضافة حاوية قادمة جديدة'
    title.font = Font(name='Calibri', size=18, bold=True, color=COLORS['text_white'])
    title.fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    title.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells('A1:B1')

    # الحقول
    fields = [
        'رقم الحاوية',
        'حجم الحاوية',
        'تاريخ الوصول',
        'الشركة',
        'البلد',
        'المحتويات',
        'الوزن (طن)',
        'حالة التصريح',
        'الشروط (إن وجدت)',
        'ملاحظات إضافية'
    ]

    row = 3
    for field in fields:
        ws.row_dimensions[row].height = 25 if field == 'ملاحظات إضافية' else 20
        label_cell = ws[f'A{row}']
        label_cell.value = field
        label_cell.font = Font(name='Calibri', size=12, bold=True, color=COLORS['text_white'])
        label_cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        label_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

        # خلية الإدخال
        input_cell = ws[f'B{row}']
        input_cell.fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type='solid')
        input_cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)

        if field == 'ملاحظات إضافية':
            ws.row_dimensions[row].height = 60

        row += 1

    # زر حفظ (تعليقات)
    save_row = row + 2
    ws.row_dimensions[save_row].height = 25
    save_cell = ws[f'B{save_row}']
    save_cell.value = '💾 اضغط Ctrl+S للحفظ'
    save_cell.font = Font(name='Calibri', size=12, bold=True, color=COLORS['text_white'])
    save_cell.fill = PatternFill(start_color=COLORS['accent'], end_color=COLORS['accent'], fill_type='solid')
    save_cell.alignment = Alignment(horizontal='center', vertical='center')

    return ws

def create_checkout_sheet(wb):
    """إنشاء ورقة تسجيل الخروج"""
    ws = wb.create_sheet('📤 تسجيل خروج', 2)
    ws.right_to_left = True
    ws.sheet_properties.tabColor = 'E67E22'

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    # العنوان
    ws.row_dimensions[1].height = 30
    title = ws['A1']
    title.value = 'تسجيل خروج وتسليم الحاوية'
    title.font = Font(name='Calibri', size=18, bold=True, color=COLORS['text_white'])
    title.fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
    title.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells('A1:B1')

    # البحث عن الحاوية
    ws.row_dimensions[3].height = 25
    search_label = ws['A3']
    search_label.value = 'بحث عن رقم الحاوية'
    search_label.font = Font(name='Calibri', size=12, bold=True, color=COLORS['text_white'])
    search_label.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')

    # الحقول
    fields = [
        ('رقم الحاوية (تلقائي)', 'A5'),
        ('حجم الحاوية (تلقائي)', 'A6'),
        ('تاريخ الوصول (تلقائي)', 'A7'),
        ('الشركة (تلقائي)', 'A8'),
        ('حالة التصريح (تلقائي)', 'A9'),
        ('الشروط إن وجدت (تلقائي)', 'A10'),
        ('تاريخ التسليم', 'A12'),
        ('ملاحظات التسليم', 'A13'),
        ('اسم المسلم', 'A14'),
    ]

    for field, cell in fields:
        ws.row_dimensions[int(cell[1:])].height = 25
        label_cell = ws[cell]
        label_cell.value = field
        label_cell.font = Font(name='Calibri', size=11, bold=True, color=COLORS['text_white'])
        label_cell.fill = PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid')

        input_cell = ws[f'B{cell[1:]}']
        input_cell.fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type='solid')
        input_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

    return ws

def create_database_sheet(wb, containers):
    """إنشاء ورقة قاعدة البيانات"""
    ws = wb.create_sheet('📁 قاعدة البيانات', 3)
    ws.right_to_left = True
    ws.sheet_properties.tabColor = '003D82'

    # رؤوس الأعمدة
    headers = [
        'رقم الحاوية', 'الحجم', 'تاريخ الوصول', 'تاريخ الموافقة',
        'الشركة', 'البلد', 'المحتويات', 'الوزن (طن)',
        'حالة التصريح', 'الشروط', 'الملاحظات', 'اسم المدخل'
    ]

    # تعيين الأعمدة
    for idx, header in enumerate(headers):
        col = get_column_letter(len(headers) - idx)
        ws.column_dimensions[col].width = 15

    # كتابة الرؤوس
    for idx, header in enumerate(headers):
        col = get_column_letter(len(headers) - idx)
        cell = ws[f'{col}1']
        style_header_cell(cell, header)

    # كتابة البيانات
    for row_idx, container in enumerate(containers, start=2):
        ws.row_dimensions[row_idx].height = 20

        data = [
            container['container_number'],
            container['size'],
            container['arrival_date'],
            container['approval_date'],
            container['company'],
            container['country'],
            container['contents'],
            container['weight_tons'],
            container['status'],
            container['conditions'],
            container['notes'],
            'النظام'  # اسم المدخل
        ]

        for col_idx, value in enumerate(data):
            col = get_column_letter(len(headers) - col_idx)
            cell = ws[f'{col}{row_idx}']
            style_data_cell(cell, value)

            # تلوين الحالات المختلفة
            if '✅' in str(value):
                cell.fill = PatternFill(start_color='D5F4E6', end_color='D5F4E6', fill_type='solid')
            elif '🔒' in str(value):
                cell.fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
            elif '⏸️' in str(value):
                cell.fill = PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid')
            else:
                alt_color = COLORS['light_gray'] if row_idx % 2 == 0 else COLORS['text_white']
                cell.fill = PatternFill(start_color=alt_color, end_color=alt_color, fill_type='solid')

    return ws

def create_reports_sheet(wb, containers):
    """إنشاء ورقة التقارير"""
    ws = wb.create_sheet('📊 التقارير', 4)
    ws.right_to_left = True
    ws.sheet_properties.tabColor = '8E44AD'

    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20

    # العنوان
    ws.row_dimensions[1].height = 30
    title = ws['A1']
    title.value = 'التقارير والإحصائيات'
    title.font = Font(name='Calibri', size=18, bold=True, color=COLORS['text_white'])
    title.fill = PatternFill(start_color='8E44AD', end_color='8E44AD', fill_type='solid')
    title.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells('A1:F1')

    # التقارير
    reports = [
        ('إجمالي الحاويات', len(containers)),
        ('حاويات 40 قدم', len([c for c in containers if c['size'] == '40 قدم'])),
        ('حاويات 20 قدم', len([c for c in containers if c['size'] == '20 قدم'])),
        ('الحاويات المحجوزة', len([c for c in containers if 'محجوزة' in c['status']])),
        ('الحاويات الممنوعة', len([c for c in containers if 'ممنوع' in c['status']])),
        ('موافقة مشروطة', len([c for c in containers if 'مشروط' in c['status']])),
        ('الحاويات المجازة', len([c for c in containers if 'يمكن' in c['status']])),
        ('إجمالي الأطنان', sum([c['weight_tons'] for c in containers])),
    ]

    row = 3
    for report_name, value in reports:
        ws.row_dimensions[row].height = 25

        name_cell = ws[f'A{row}']
        name_cell.value = report_name
        name_cell.font = Font(name='Calibri', size=12, bold=True, color=COLORS['text_white'])
        name_cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')

        value_cell = ws[f'B{row}']
        value_cell.value = value
        value_cell.font = Font(name='Calibri', size=14, bold=True, color=COLORS['accent'])
        value_cell.fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type='solid')
        value_cell.alignment = Alignment(horizontal='center', vertical='center')

        row += 1

    # إحصائيات حسب البلد
    row += 2
    ws.row_dimensions[row].height = 25
    country_title = ws[f'A{row}']
    country_title.value = 'الحاويات حسب البلد'
    country_title.font = Font(name='Calibri', size=12, bold=True, color=COLORS['text_white'])
    country_title.fill = PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid')
    ws.merge_cells(f'A{row}:B{row}')

    row += 1
    for country in set([c['country'] for c in containers]):
        count = len([c for c in containers if c['country'] == country])
        ws.row_dimensions[row].height = 20

        ws[f'A{row}'].value = country
        ws[f'B{row}'].value = count
        ws[f'A{row}'].font = Font(name='Calibri', size=11)
        ws[f'B{row}'].font = Font(name='Calibri', size=11)
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        row += 1

    return ws

def create_certificates_sheet(wb):
    """إنشاء ورقة الشهادات"""
    ws = wb.create_sheet('🖨️ الشهادات', 5)
    ws.right_to_left = True
    ws.sheet_properties.tabColor = 'C0392B'

    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 20

    # العنوان
    ws.row_dimensions[1].height = 30
    title = ws['A1']
    title.value = 'شهادات الاستقبال والتسليم'
    title.font = Font(name='Calibri', size=18, bold=True, color=COLORS['text_white'])
    title.fill = PatternFill(start_color=COLORS['danger'], end_color=COLORS['danger'], fill_type='solid')
    title.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells('A1:D1')

    # شهادة الاستقبال
    row = 3
    ws.row_dimensions[row].height = 25
    receipt_title = ws[f'A{row}']
    receipt_title.value = 'شهادة استقبال الحاوية'
    receipt_title.font = Font(name='Calibri', size=14, bold=True, color=COLORS['text_white'])
    receipt_title.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    ws.merge_cells(f'A{row}:D{row}')

    row += 2
    receipt_fields = [
        'رقم الشهادة',
        'رقم الحاوية',
        'تاريخ الاستقبال',
        'الموقع (الميناء)',
        'المسؤول عن الاستقبال',
        'التوقيع والختم',
    ]

    for field in receipt_fields:
        ws.row_dimensions[row].height = 20
        ws[f'A{row}'].value = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color=COLORS['text_white'])
        ws[f'A{row}'].fill = PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid')

        input_cell = ws[f'B{row}']
        input_cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        input_cell.border = Border(bottom=Side(style='thin', color='000000'))

        row += 1

    # شهادة التسليم
    row += 2
    ws.row_dimensions[row].height = 25
    delivery_title = ws[f'A{row}']
    delivery_title.value = 'شهادة تسليم الحاوية'
    delivery_title.font = Font(name='Calibri', size=14, bold=True, color=COLORS['text_white'])
    delivery_title.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    ws.merge_cells(f'A{row}:D{row}')

    row += 2
    delivery_fields = [
        'رقم الشهادة',
        'رقم الحاوية',
        'تاريخ التسليم',
        'اسم المستقبل',
        'الشركة',
        'التوقيع والختم',
    ]

    for field in delivery_fields:
        ws.row_dimensions[row].height = 20
        ws[f'A{row}'].value = field
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color=COLORS['text_white'])
        ws[f'A{row}'].fill = PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid')

        input_cell = ws[f'B{row}']
        input_cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        input_cell.border = Border(bottom=Side(style='thin', color='000000'))

        row += 1

    return ws

def create_settings_sheet(wb):
    """إنشاء ورقة الإعدادات (مخفية)"""
    ws = wb.create_sheet('⚙️ الإعدادات', 6)
    ws.right_to_left = True
    ws.sheet_properties.tabColor = '95A5A6'
    ws.sheet_state = 'hidden'  # إخفاء الورقة

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40

    # قائمة الأحجام
    ws['A1'].value = 'أحجام الحاويات'
    ws['A1'].font = Font(bold=True)
    sizes = ['20 قدم', '40 قدم', 'High Cube']
    for idx, size in enumerate(sizes, start=2):
        ws[f'A{idx}'].value = size

    # قائمة البلدان
    ws['D1'].value = 'البلدان'
    ws['D1'].font = Font(bold=True)
    for idx, country in enumerate(COUNTRIES, start=2):
        ws[f'D{idx}'].value = country

    # قائمة الشركات
    ws['F1'].value = 'الشركات'
    ws['F1'].font = Font(bold=True)
    for idx, company in enumerate(COMPANIES, start=2):
        ws[f'F{idx}'].value = company

    # حالات التصريح
    ws['H1'].value = 'حالات التصريح'
    ws['H1'].font = Font(bold=True)
    for idx, status in enumerate(STATUS_OPTIONS, start=2):
        ws[f'H{idx}'].value = status

    return ws

# ============================================================================
# المجمع الرئيسي
# ============================================================================

def main():
    """البرنامج الرئيسي - إنشاء ملف Excel الكامل"""
    print('🚀 بدء إنشاء نظام إدارة الحاويات...')
    print()

    # 1. إنشاء Workbook
    print('📝 إنشاء ملف Excel...')
    wb = create_workbook()
    wb.title = 'نظام إدارة الحاويات'

    # 2. توليد البيانات التجريبية
    print('📦 توليد 15 حاوية تجريبية...')
    containers = generate_sample_containers()

    # 3. إنشاء الأوراق
    print('📊 إنشاء لوحة التحكم...')
    create_dashboard(wb, containers)

    print('📥 إنشاء ورقة إضافة الحاوية...')
    create_add_container_sheet(wb)

    print('📤 إنشاء ورقة تسجيل الخروج...')
    create_checkout_sheet(wb)

    print('📁 إنشاء قاعدة البيانات...')
    create_database_sheet(wb, containers)

    print('📊 إنشاء ورقة التقارير...')
    create_reports_sheet(wb, containers)

    print('🖨️ إنشاء ورقة الشهادات...')
    create_certificates_sheet(wb)

    print('⚙️ إنشاء ورقة الإعدادات...')
    create_settings_sheet(wb)

    # 4. حفظ الملف
    output_path = 'نظام_إدارة_الحاويات.xlsx'
    print()
    print(f'💾 حفظ الملف: {output_path}')
    wb.save(output_path)

    print()
    print('✅ تم إنشاء نظام إدارة الحاويات بنجاح!')
    print(f'📍 الملف: {output_path}')
    print()
    print('📊 الأوراق المُنشأة:')
    print('  1. 📊 لوحة التحكم')
    print('  2. 📥 إضافة حاوية قادمة')
    print('  3. 📤 تسجيل خروج')
    print('  4. 📁 قاعدة البيانات')
    print('  5. 📊 التقارير')
    print('  6. 🖨️ الشهادات')
    print('  7. ⚙️ الإعدادات (مخفية)')
    print()
    print('🎨 تصميم احترافي RTL عربي')
    print('💾 البيانات: 15 حاوية تجريبية')
    print('✨ جاهز للاستخدام الفوري!')

if __name__ == '__main__':
    main()
